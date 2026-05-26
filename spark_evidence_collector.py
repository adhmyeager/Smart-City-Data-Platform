"""
spark_evidence_collector.py  (FIXED VERSION)
=============================================
يقرأ الـ output اللي Spark حفظه ويعمل Excel report كدليل إن Spark اشتغل.

المشاكل اللي اتصلحت:
  1. auto_offset_reset="earliest"  بدل latest  → بيقرأ الـ records القديمة
  2. consumer_timeout_ms=12000     (12 ثانية بدل 5) → وقت أكتر للـ consumer
  3. بيعمل poll صح قبل ما يسيك
  4. لو Kafka فاضية أو مش شغالة → بيقولك السبب بالظبط

شغّله:
    py -3.11 spark_evidence_collector.py
"""

import subprocess, json, datetime, os, sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

try:
    from kafka import KafkaConsumer, TopicPartition
    from kafka.errors import KafkaError
    KAFKA_AVAILABLE = True
except ImportError:
    KAFKA_AVAILABLE = False

# ── Colors ────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GOLD       = "F4B942"
GREEN      = "70AD47"
RED        = "FF0000"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

# ── Helpers ───────────────────────────────────────────────────

def cell_style(ws, cell_ref, value, bold=False, bg=None, fg=WHITE,
               align="center", size=11, border=False):
    c = ws[cell_ref]
    c.value = value
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        thin = Side(style="thin", color="BFBFBF")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_row(ws, row, cols_vals, bg=DARK_BLUE):
    for col, val in cols_vals:
        cell_style(ws, f"{col}{row}", val, bold=True, bg=bg, fg=WHITE,
                   align="center", size=11, border=True)

def data_row(ws, row, cols_vals, bg=None):
    actual_bg = bg if bg else (LIGHT_GRAY if row % 2 == 0 else WHITE)
    for col, val in cols_vals:
        cell_style(ws, f"{col}{row}", val, bg=actual_bg, fg="000000",
                   align="left", border=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def merge_title(ws, cell_range, value, bg=DARK_BLUE, size=14):
    ws.merge_cells(cell_range)
    start = cell_range.split(":")[0]
    c = ws[start]
    c.value = value
    c.font = Font(bold=True, color=WHITE, size=size, name="Calibri")
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int(''.join(filter(str.isdigit, start)))].height = 32

# ── Data Collection ───────────────────────────────────────────

def collect_kafka_samples(n=30):
    """
    FIX: يستخدم earliest + manual seek عشان يقرأ آخر n records حتى لو
    الـ producer مش شغال دلوقتي.
    """
    if not KAFKA_AVAILABLE:
        print("  [Kafka] kafka-python مش متنصّب — شغّل: pip install kafka-python")
        return []

    servers = ["127.0.0.1:19092", "localhost:19092"]
    for server in servers:
        try:
            print(f"  [Kafka] بتجرب {server} ...")
            consumer = KafkaConsumer(
                bootstrap_servers=server,
                auto_offset_reset="earliest",       # ← FIX الأساسي
                consumer_timeout_ms=12000,           # ← 12 ثانية
                value_deserializer=lambda m: json.loads(m.decode("utf-8")),
                group_id=None,
                request_timeout_ms=15000,
                session_timeout_ms=10000,
            )

            # Manual assignment + seek to end - n
            tp = TopicPartition("vehicle-telemetry", 0)
            consumer.assign([tp])
            consumer.poll(timeout_ms=3000)          # trigger metadata fetch

            end_offset = consumer.end_offsets([tp])[tp]
            begin_offset = consumer.beginning_offsets([tp])[tp]

            if end_offset == 0:
                print(f"  [Kafka] Topic فاضي — شغّل kafka_producer.py الأول")
                consumer.close()
                return []

            # Seek to last n records
            seek_to = max(begin_offset, end_offset - n)
            consumer.seek(tp, seek_to)
            print(f"  [Kafka] Offset range: {begin_offset}→{end_offset}, قراءة من {seek_to}")

            records = []
            for msg in consumer:
                records.append(msg.value)
                if len(records) >= n:
                    break

            consumer.close()
            print(f"  [Kafka] ✅ جاب {len(records)} records")
            return records

        except Exception as e:
            print(f"  [Kafka] {server} فشل: {e}")
            continue

    print("  [Kafka] ❌ مش قادر يتصل — تأكد Docker شغال وكافكا على بورت 19092")
    return []


def collect_spark_parquet():
    """Check parquet files inside Spark container."""
    results = {}
    paths = {
        "bronze/telemetry":  "/tmp/spark_output/bronze/telemetry",
        "silver/telemetry":  "/tmp/spark_output/silver/telemetry",
        "gold/vehicle_kpis": "/tmp/spark_output/gold/vehicle_kpis",
        "alerts":            "/tmp/spark_output/alerts",
    }
    for label, path in paths.items():
        try:
            out = subprocess.run(
                ["docker", "exec", "sc_spark_master", "find", path,
                 "-name", "*.parquet", "-type", "f"],
                capture_output=True, text=True, timeout=10
            )
            files = [f for f in out.stdout.strip().split("\n") if f]
            results[label] = {"files": len(files), "path": path, "found": len(files) > 0}
        except Exception as e:
            results[label] = {"files": 0, "path": path, "found": False, "error": str(e)}
    return results


def get_spark_app_info():
    """Get Spark application info from REST API."""
    try:
        import urllib.request
        url = "http://localhost:8082/api/v1/applications"
        with urllib.request.urlopen(url, timeout=5) as r:
            return json.loads(r.read())
    except Exception as e:
        print(f"  [Spark API] {e}")
        return []


def get_kafka_topic_stats():
    """Get topic stats: message count per topic."""
    stats = {}
    topics = ["vehicle-telemetry", "weather-data", "traffic-events"]
    if not KAFKA_AVAILABLE:
        return stats
    try:
        from kafka import TopicPartition as TP
        consumer = KafkaConsumer(
            bootstrap_servers="127.0.0.1:19092",
            consumer_timeout_ms=5000,
            group_id=None,
            request_timeout_ms=10000,
        )
        for topic in topics:
            try:
                partitions = consumer.partitions_for_topic(topic) or {0}
                tps = [TP(topic, p) for p in partitions]
                ends   = consumer.end_offsets(tps)
                starts = consumer.beginning_offsets(tps)
                total  = sum(ends[tp] - starts[tp] for tp in tps)
                stats[topic] = total
            except Exception:
                stats[topic] = "?"
        consumer.close()
    except Exception as e:
        print(f"  [Topic stats] {e}")
    return stats

# ── Excel Builder ─────────────────────────────────────────────

def build_excel(kafka_records, spark_files, spark_apps, topic_stats):
    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════
    # Sheet 1 — Dashboard
    # ═══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 10

    merge_title(ws1, "B2:H3",
                "CairoFlow Smart City — Spark Evidence Report", size=16)

    ws1["B4"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    ws1["B4"].font = Font(color="888888", size=9, italic=True)

    ws1.row_dimensions[6].height = 14
    merge_title(ws1, "B7:C8",
                f"Kafka Records\n{len(kafka_records)} captured", bg=MED_BLUE, size=12)

    spark_total = sum(v["files"] for v in spark_files.values())
    merge_title(ws1, "D7:E8",
                f"Parquet Files\n{spark_total} written", bg=GREEN, size=12)

    app_count = len(spark_apps)
    merge_title(ws1, "F7:G8",
                f"Spark Apps\n{app_count} detected",
                bg=GOLD if app_count else RED, size=12)

    status = "PIPELINE WORKING" if (len(kafka_records) > 0 or spark_total > 0) else "NO DATA YET — Run kafka_producer.py first"
    status_color = GREEN if "WORKING" in status else GOLD
    merge_title(ws1, "B10:G11", status, bg=status_color, size=14)

    # Kafka topic stats
    ws1.row_dimensions[13].height = 8
    merge_title(ws1, "B14:G14", "Kafka Topics — Message Count", bg=MED_BLUE, size=11)
    header_row(ws1, 15, [("B","Topic"),("C","Messages"),("D","Status")])
    for i, (topic, count) in enumerate(topic_stats.items(), start=16):
        ok = isinstance(count, int) and count > 0
        data_row(ws1, i, [
            ("B", topic),
            ("C", count),
            ("D", "Active" if ok else "Empty / Unknown"),
        ])

    # Spark layers
    ws1.row_dimensions[20].height = 8
    merge_title(ws1, "B21:G21", "Spark Output Layers", bg=MED_BLUE, size=11)
    header_row(ws1, 22, [("B","Layer"),("C","Path"),("D","Parquet Files"),("E","Status"),("F","Purpose")])

    layer_info = {
        "bronze/telemetry":  ("Bronze", "Raw ingest — data as-is from Kafka"),
        "silver/telemetry":  ("Silver", "Cleaned — filtered & classified"),
        "gold/vehicle_kpis": ("Gold",   "Aggregated KPIs per 30s window"),
        "alerts":            ("Alerts", "Real-time anomaly detection"),
    }
    for i, (key, info) in enumerate(layer_info.items(), start=23):
        v = spark_files.get(key, {})
        found = v.get("found", False)
        data_row(ws1, i, [
            ("B", info[0]),
            ("C", v.get("path", "")),
            ("D", v.get("files", 0)),
            ("E", "Found" if found else "Not yet — run Spark job"),
            ("F", info[1]),
        ])

    set_col_widths(ws1, {"A":3,"B":18,"C":40,"D":15,"E":28,"F":40,"G":5,"H":5})

    # ═══════════════════════════════════════════════════
    # Sheet 2 — Vehicle Telemetry (Kafka samples)
    # ═══════════════════════════════════════════════════
    ws2 = wb.create_sheet("Vehicle Telemetry")
    ws2.sheet_view.showGridLines = False
    merge_title(ws2, "A1:P1", "Vehicle Telemetry — Kafka Samples (last 30 records)", bg=DARK_BLUE, size=13)

    cols = [
        ("A","vehicle_id"),("B","vehicle_type"),("C","route_name"),
        ("D","speed_kmh"),("E","rpm"),("F","gear"),
        ("G","engine_temp_c"),("H","fuel_level_pct"),("I","fuel_rate_l100km"),
        ("J","latitude"),("K","longitude"),
        ("L","road_type"),("M","road_event"),("N","traffic_density"),
        ("O","timestamp_iso"),("P","source"),
    ]
    header_row(ws2, 2, cols, bg=MED_BLUE)

    for r, rec in enumerate(kafka_records, start=3):
        row_bg = None
        if (rec.get("speed_kmh", 0) > 120 or
                rec.get("engine_temp_c", 0) > 105 or
                rec.get("fuel_level_pct", 100) < 10):
            row_bg = "FFE0E0"
        data_row(ws2, r, [
            ("A", rec.get("vehicle_id","")),
            ("B", rec.get("vehicle_type","")),
            ("C", rec.get("route_name","")),
            ("D", rec.get("speed_kmh","")),
            ("E", rec.get("rpm","")),
            ("F", rec.get("gear","")),
            ("G", rec.get("engine_temp_c","")),
            ("H", rec.get("fuel_level_pct","")),
            ("I", rec.get("fuel_rate_l100km","")),
            ("J", rec.get("latitude","")),
            ("K", rec.get("longitude","")),
            ("L", rec.get("road_type","")),
            ("M", rec.get("road_event","")),
            ("N", rec.get("traffic_density","")),
            ("O", rec.get("timestamp_iso","")),
            ("P", "Kafka OK"),
        ], bg=row_bg)

    if not kafka_records:
        c = ws2["A3"]
        c.value = "No Kafka records — make sure kafka_producer.py is running, then rerun this script"
        c.font = Font(color=RED, bold=True)

    set_col_widths(ws2, {
        "A":12,"B":10,"C":28,"D":11,"E":8,"F":6,
        "G":14,"H":15,"I":16,"J":12,"K":12,
        "L":10,"M":18,"N":15,"O":26,"P":10,
    })

    # ═══════════════════════════════════════════════════
    # Sheet 3 — Spark Apps
    # ═══════════════════════════════════════════════════
    ws3 = wb.create_sheet("Spark Apps")
    ws3.sheet_view.showGridLines = False
    merge_title(ws3, "A1:F1", "Spark Applications History", bg=DARK_BLUE, size=13)
    header_row(ws3, 2, [
        ("A","App ID"),("B","Name"),("C","State"),
        ("D","Start Time"),("E","Duration (s)"),("F","User"),
    ], bg=MED_BLUE)

    if spark_apps:
        for r, app in enumerate(spark_apps, start=3):
            attempts = app.get("attempts", [{}])
            last = attempts[-1] if attempts else {}
            duration = last.get("duration", 0) // 1000
            state = "RUNNING" if last.get("completed") == False else "FINISHED"
            data_row(ws3, r, [
                ("A", app.get("id","")),
                ("B", app.get("name","")),
                ("C", state),
                ("D", last.get("startTime","")),
                ("E", duration),
                ("F", last.get("sparkUser","spark")),
            ])
    else:
        c = ws3["A3"]
        c.value = "Could not reach Spark REST API — check http://localhost:8082"
        c.font = Font(color=RED, bold=True)

    set_col_widths(ws3, {"A":30,"B":35,"C":12,"D":30,"E":14,"F":10})

    # ═══════════════════════════════════════════════════
    # Sheet 4 — Architecture
    # ═══════════════════════════════════════════════════
    ws4 = wb.create_sheet("Architecture")
    ws4.sheet_view.showGridLines = False
    merge_title(ws4, "A1:D1", "Pipeline Architecture", bg=DARK_BLUE, size=13)

    arch = [
        ("Source",    "Python Vehicle Simulator",   "kafka_producer.py",            "5 cars sending data every second"),
        ("Ingestion", "Apache Kafka + Zookeeper",    "localhost:19092 / port 2181",  "Message broker — receives & stores events"),
        ("Streaming", "Apache Spark Structured",     "spark://sc_spark_master:7077", "Processes stream in real-time"),
        ("Bronze",    "Raw Parquet Storage",          "/tmp/spark_output/bronze",    "Raw data — no transformation"),
        ("Silver",    "Cleaned Parquet Storage",      "/tmp/spark_output/silver",    "After filtering & classification"),
        ("Gold",      "Aggregated KPIs",              "/tmp/spark_output/gold",      "Window aggregations every 30s"),
        ("Alerts",    "Anomaly Detection",            "/tmp/spark_output/alerts",    "Speed/Temp/Fuel alerts in real-time"),
        ("Monitor",   "Kafka UI",                     "localhost:8080",               "View topics & messages"),
        ("Monitor",   "Spark UI",                     "localhost:8082",               "View jobs & executors"),
    ]

    header_row(ws4, 2, [("A","Layer"),("B","Component"),("C","Endpoint/Path"),("D","Description")], bg=MED_BLUE)
    for r, row in enumerate(arch, start=3):
        data_row(ws4, r, [("A",row[0]),("B",row[1]),("C",row[2]),("D",row[3])])

    set_col_widths(ws4, {"A":14,"B":28,"C":36,"D":45})

    return wb


# ── Main ──────────────────────────────────────────────────────

def main():
    print("\n══════════════════════════════════════════════")
    print("  CairoFlow — Spark Evidence Collector (FIXED)")
    print("══════════════════════════════════════════════\n")

    print("Collecting Kafka samples...")
    kafka_records = collect_kafka_samples(n=30)
    print(f"  Got {len(kafka_records)} records\n")

    print("Getting topic stats...")
    topic_stats = get_kafka_topic_stats()
    for t, c in topic_stats.items():
        print(f"  {t}: {c} messages")
    print()

    print("Checking Spark Parquet output...")
    spark_files = collect_spark_parquet()
    for k, v in spark_files.items():
        print(f"  {k}: {v['files']} files {'OK' if v['found'] else 'not yet'}")
    print()

    print("Checking Spark REST API...")
    spark_apps = get_spark_app_info()
    print(f"  Found {len(spark_apps)} application(s)\n")

    print("Building Excel report...")
    wb = build_excel(kafka_records, spark_files, spark_apps, topic_stats)

    out = "spark_report.xlsx"
    wb.save(out)
    print(f"\nReport saved: {out}")
    print(f"Sheets: Dashboard | Vehicle Telemetry | Spark Apps | Architecture\n")


if __name__ == "__main__":
    main()
